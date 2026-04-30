"""Build BALANCE.xlsx — consolidated balance reference extracted from GAME.md.

All numbers are pulled directly from GAME.md Part I. This file is a designer-facing
spot-check tool: tweak a number, see the derived rates/margins/capacities update.
Hardcoded inputs are blue; formulas are black; cross-sheet pulls are green.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
INPUT_BLUE = Font(name=FONT, color="0000FF")
FORMULA_BLACK = Font(name=FONT, color="000000")
LINK_GREEN = Font(name=FONT, color="008000")
HEADER = Font(name=FONT, bold=True, color="FFFFFF")
SECTION = Font(name=FONT, bold=True, size=14)
NOTE = Font(name=FONT, italic=True, color="606060", size=9)
HEADER_FILL = PatternFill("solid", start_color="1F3864")
SECTION_FILL = PatternFill("solid", start_color="D9E1F2")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
wb.remove(wb.active)


def header_row(sheet, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = sheet.cell(row=row, column=i, value=h)
        c.font = HEADER
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.row_dimensions[row].height = 28


def section_row(sheet, row, ncols, label):
    sheet.cell(row=row, column=1, value=label).font = SECTION
    sheet.cell(row=row, column=1).fill = SECTION_FILL
    for i in range(2, ncols + 1):
        sheet.cell(row=row, column=i).fill = SECTION_FILL
    sheet.row_dimensions[row].height = 20


def write_row(sheet, row, values, font_map=None, fills=None):
    """font_map: dict of col_idx (1-based) -> Font. Default INPUT_BLUE for non-text, black for text."""
    for i, v in enumerate(values, 1):
        c = sheet.cell(row=row, column=i, value=v)
        c.border = BORDER
        if font_map and i in font_map:
            c.font = font_map[i]
        else:
            if isinstance(v, str) and (v.startswith("=") or "!" in (v or "")):
                c.font = FORMULA_BLACK if v.startswith("=") and "!" not in v else LINK_GREEN
            elif isinstance(v, (int, float)):
                c.font = INPUT_BLUE
            else:
                c.font = Font(name=FONT)
        if fills and i in fills:
            c.fill = fills[i]


# =============================================================================
# 1. INDEX
# =============================================================================
ws = wb.create_sheet("Index")
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 78
ws["A1"] = "VoidYield 2 — Balance Sheet"
ws["A1"].font = Font(name=FONT, bold=True, size=18)
ws["A2"] = "Source: GAME.md (Part I) — extracted 2026-04-29 after Round 12 (R72–R77)"
ws["A2"].font = NOTE
ws["A3"] = "All values are placeholders for Stage 3 playtest tuning. Edit blue cells; black cells recalculate."
ws["A3"].font = NOTE

ws["A5"] = "Sheet"
ws["B5"] = "Contents"
for c in (ws["A5"], ws["B5"]):
    c.font = HEADER
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")

index_rows = [
    ("Tier_Ladder", "8 named tiers (T0 Wildcatter → T7 System Corporation), gate conditions, headline capabilities"),
    ("Resources", "All 27 resources T0–T3: tier, class, cargo, Earth buy/sell prices + computed margin/margin%"),
    ("Recipes", "All 31 recipes T0–T3: cycle time, inputs, outputs, derived rate/min, value-add/cycle"),
    ("Adjacency_Pairs", "All 22 pair-type bonuses T0–T3 (R72): placer + pair + bonus magnitude"),
    ("People_Costs", "All buildings + people_cost (R73). Sanity-check: total starter T0 chain cost vs. 8-person crew"),
    ("Pop_Tiers", "Six pop tiers + multipliers + settle-in windows + continuous needs + growth-tier bundles"),
    ("Storage_Caps", "Silo/Tank/Cryo capacities by tier transition"),
    ("Fuel_Costs", "Per-route base fuel costs (R74); window multipliers; computed leg costs by hull"),
    ("Ships", "6 hull catalog T0–T2 (capacity, speed, fuel/route multiplier, Earth buy)"),
    ("Tier_Gates", "Primary + alternate fulfillment for T0→T1 through T3→T4 (R76)"),
    ("Action_Durations", "Build/transit/cycle times across the game"),
    ("Earth_Prefab_Kits", "T1–T3 Earth-bought prefab kits (1-of-1 per tier)"),
]
for i, (sheet_name, desc) in enumerate(index_rows, start=6):
    ws.cell(row=i, column=1, value=sheet_name).font = Font(name=FONT, bold=True, color="0563C1", underline="single")
    ws.cell(row=i, column=1).hyperlink = f"#{sheet_name}!A1"
    ws.cell(row=i, column=2, value=desc).font = Font(name=FONT)
    ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 28

ws["A20"] = "Color Legend"
ws["A20"].font = Font(name=FONT, bold=True)
ws["A21"] = "Blue cells"
ws["A21"].font = INPUT_BLUE
ws["B21"] = "Hardcoded inputs from GAME.md — change to test what-if scenarios."
ws["A22"] = "Black cells"
ws["A22"].font = FORMULA_BLACK
ws["B22"] = "Computed formulas — recalculate when inputs change."
ws["A23"] = "Green cells"
ws["A23"].font = LINK_GREEN
ws["B23"] = "Cross-sheet links — pull values from other sheets."

# =============================================================================
# 2. TIER LADDER
# =============================================================================
ws = wb.create_sheet("Tier_Ladder")
header_row(ws, 1, ["Tier", "Name", "Region Unlocked", "Gate to Next Tier", "Headline New Capability"], [6, 22, 38, 50, 50])
ladder = [
    ("T0", "Wildcatter", "Earth orbit + 1 NEA", "Sell N metals to Earth", "Ship surveying, basic mining"),
    ("T1", "Lunar Foothold", "Moon orbit + lunar surface site", "First habitat reaches Pop 50", "First colony, life-support imports"),
    ("T2", "NEA Industry", "NEA cluster (3–5 asteroids)", "Local oxygen + water production", "Tankers, fluid/gas cargo class"),
    ("T3", "Cislunar Network", "Rare-trace NEAs (~15–20% of NEAs)", "Mars shipyard built; Networked-tier colony", "Maintain-stock + surplus-export rules; Networked pop tier"),
    ("T4", "Martian Reach", "Mars orbit + Phobos/Deimos", "(deferred to T4 drill — Pending #39)", "Local shipbuilding, advanced alloys"),
    ("T5", "Belt Operations", "Main belt mining hub", "3 tier-3 colonies sustaining", "Bulk haulers, advanced refining"),
    ("T6", "Jovian Frontier", "Jovian moons (Europa/Ganymede focus)", "Helium-3 / heavy isotope chain", "Long-range drives, ice giants"),
    ("T7", "System Corporation", "Saturn + outer system probes", "Endgame milestone bundle (5–10h destination)", "Prestige unlock"),
]
for i, row in enumerate(ladder, start=2):
    write_row(ws, i, row, font_map={k: Font(name=FONT) for k in range(1, 6)})
    if i % 2 == 0:
        for j in range(1, 6):
            ws.cell(row=i, column=j).fill = ALT_FILL
ws.cell(row=11, column=1, value="Note: gate conditions are 'content gates, not paywalls' — every gate is something the player produces or builds.").font = NOTE

# =============================================================================
# 3. RESOURCES
# =============================================================================
ws = wb.create_sheet("Resources")
header_row(ws, 1, ["#", "Resource", "Tier", "Class", "Cargo", "Earth Buy", "Earth Sell", "Margin (Buy-Sell)", "Margin %", "Notes"],
           [4, 26, 6, 14, 12, 11, 11, 17, 12, 35])

resources = [
    (1, "Iron Ore", "T0", "Raw", "Solid", 3, 1, ""),
    (2, "Water Ice", "T0", "Raw", "Solid", 4, 2, ""),
    (3, "Refined Metal", "T0", "Intermediate", "Solid", 18, 12, ""),
    (4, "Hydrogen Fuel", "T0", "Intermediate", "Fluid/Gas", 8, 5, "v1 fuel through T2"),
    (5, "Oxygen", "T0*", "Intermediate", "Fluid/Gas", 6, 3, "Electrolysis byproduct; load-bearing at T1"),
    (6, "Lunar Regolith", "T1", "Raw", "Solid", None, 2, "No Earth buy"),
    (7, "Aluminum", "T1", "Intermediate", "Solid", 22, 15, ""),
    (8, "Construction Materials", "T1", "Finished", "Solid", 60, 45, ""),
    (9, "Food Pack", "T1", "Finished", "Solid", 25, 18, ""),
    (10, "Habitat Module", "T1", "Finished", "Solid", 180, 130, "Earth Prefab Kit at T1"),
    (11, "Nickel Ore", "T2", "Raw", "Solid", 5, 2, ""),
    (12, "Carbonaceous Ore", "T2", "Raw", "Solid", 4, 2, ""),
    (13, "Silicates", "T2", "Raw", "Solid", 4, 2, ""),
    (14, "Pressure Valves", "T2", "Finished", "Solid", 90, 65, ""),
    (15, "Habitat Glass", "T2", "Finished", "Solid", 75, 55, ""),
    (16, "Carbon Mesh", "T2", "Intermediate", "Solid", 40, 28, ""),
    (17, "Textiles", "T2", "Finished", "Solid", 70, 50, ""),
    (18, "Furnishings", "T2", "Finished", "Solid", 110, 80, ""),
    (19, "Spirits", "T2", "Finished", "Fluid/Gas", 95, 70, ""),
    (20, "Hydroponic Yield", "T2", "Intermediate", "Solid", None, 12, "No Earth buy"),
    (21, "Rare Trace Elements", "T3", "Raw", "Solid", 30, 18, "Rare-trace NEAs only (~15–20%)"),
    (22, "Ammonia", "T3", "Raw", "Fluid/Gas", 12, 7, "Earth buy only until T6"),
    (23, "Logic Boards", "T3", "Intermediate", "Solid", 80, 55, ""),
    (24, "Refined Hydrogen", "T3", "Intermediate", "Fluid/Gas", 25, 16, "T3+ advanced-hull fuel grade"),
    (25, "Comms Module", "T3", "Finished", "Solid", 250, 180, ""),
    (26, "Sensor Array", "T3", "Finished", "Solid", 220, 160, ""),
    (27, "Maintenance Kit", "T3", "Finished", "Solid", 180, 130, ""),
]
for i, (num, name, tier, cls, cargo, buy, sell, notes) in enumerate(resources, start=2):
    ws.cell(row=i, column=1, value=num).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=name).font = Font(name=FONT)
    ws.cell(row=i, column=3, value=tier).font = Font(name=FONT)
    ws.cell(row=i, column=4, value=cls).font = Font(name=FONT)
    ws.cell(row=i, column=5, value=cargo).font = Font(name=FONT)
    ws.cell(row=i, column=6, value=buy).font = INPUT_BLUE
    ws.cell(row=i, column=7, value=sell).font = INPUT_BLUE
    if buy is not None and sell is not None:
        ws.cell(row=i, column=8, value=f"=F{i}-G{i}").font = FORMULA_BLACK
        ws.cell(row=i, column=9, value=f"=IFERROR((F{i}-G{i})/F{i},0)").font = FORMULA_BLACK
        ws.cell(row=i, column=9).number_format = "0.0%"
    else:
        ws.cell(row=i, column=8, value="—").font = Font(name=FONT)
        ws.cell(row=i, column=9, value="—").font = Font(name=FONT)
    ws.cell(row=i, column=10, value=notes).font = Font(name=FONT)
    for j in range(1, 11):
        ws.cell(row=i, column=j).border = BORDER
        if i % 2 == 0:
            existing_fill = ws.cell(row=i, column=j).fill
            if existing_fill.start_color.rgb in (None, "00000000"):
                ws.cell(row=i, column=j).fill = ALT_FILL
ws.cell(row=len(resources) + 3, column=1,
        value="Margin = Earth Buy − Earth Sell (the spread Earth captures). Margin % = margin / Buy.").font = NOTE
ws.cell(row=len(resources) + 4, column=1, value="* Oxygen: T0 byproduct only at low margin; load-bearing for life support from T1 onward.").font = NOTE

# =============================================================================
# 4. RECIPES (with rate/min and value-add formulas)
# =============================================================================
ws = wb.create_sheet("Recipes")
header_row(ws, 1, ["Tier", "Building", "Cycle (s)", "Inputs", "Outputs", "Output Qty", "Output Resource",
                   "Cycles/min", "Output rate/min", "Output sell price (Earth)", "Output value/min ($)", "Notes"],
           [5, 28, 9, 38, 30, 9, 22, 9, 14, 12, 14, 30])

# (tier, name, cycle_s, inputs_text, outputs_text, output_qty, output_resource_for_pricing, notes)
recipes = [
    ("T0", "Small Mine", 30, "—", "10 Iron Ore", 10, "Iron Ore", ""),
    ("T0", "Ice Mine", 40, "—", "8 Water Ice", 8, "Water Ice", ""),
    ("T0", "Smelter", 45, "5 Iron Ore", "2 Refined Metal", 2, "Refined Metal", "Signature pair: +30% near Mine"),
    ("T0", "Electrolyzer", 60, "4 Water Ice", "3 Hydrogen Fuel + 1 Oxygen", 3, "Hydrogen Fuel", "Output also produces 1 Oxygen"),
    ("T0", "Probe Bay", 0, "—", "passive surveying", 0, "", "Not a recipe; survey time = 4 min"),
    ("T0", "Silo", 0, "—", "+300 solids capacity", 0, "", "Storage, neutral"),
    ("T0", "Tank", 0, "—", "+180 fluids/gases capacity", 0, "", "Storage, neutral"),
    ("T1", "Lunar Surface Mine", 50, "—", "6 Lunar Regolith", 6, "Lunar Regolith", ""),
    ("T1", "Refinery (Aluminum)", 70, "3 Lunar Regolith", "2 Aluminum", 2, "Aluminum", "Signature: +30% near Lunar Surface Mine"),
    ("T1", "Construction Yard", 90, "2 Refined Metal + 2 Aluminum", "1 Construction Materials", 1, "Construction Materials", ""),
    ("T1", "Habitat Assembler", 480, "6 Construction Materials", "1 Habitat Module", 1, "Habitat Module", "8 min/cycle — slow on purpose"),
    ("T1", "Greenhouse (small)", 60, "2 Water Ice", "2 Food Pack", 2, "Food Pack", ""),
    ("T1", "Life Support — Water", 0, "1 Water Ice / pop / 8 min", "—", 0, "", "Continuous draw"),
    ("T1", "Life Support — Oxygen", 0, "1 Oxygen / pop / 6 min", "—", 0, "", "Continuous draw"),
    ("T1", "Life Support — Food", 0, "1 Food Pack / pop / 12 min", "—", 0, "", "Continuous draw"),
    ("T2", "NEA Mine (Nickel)", 45, "—", "6 Nickel Ore", 6, "Nickel Ore", ""),
    ("T2", "NEA Mine (Carbon)", 45, "—", "6 Carbonaceous Ore", 6, "Carbonaceous Ore", ""),
    ("T2", "NEA Mine (Silicates)", 50, "—", "7 Silicates", 7, "Silicates", ""),
    ("T2", "Hydroponics Bay", 60, "2 Water Ice", "2 Hydroponic Yield", 2, "Hydroponic Yield", ""),
    ("T2", "Hydroponic Greenhouse", 45, "2 Hydroponic Yield", "4 Food Pack", 4, "Food Pack", "T2 upgrade replaces small greenhouse"),
    ("T2", "Glass Furnace", 90, "4 Silicates + 1 Aluminum", "2 Habitat Glass", 2, "Habitat Glass", "Signature: +30% near Silicates Mine"),
    ("T2", "Carbon Mill", 75, "3 Carbonaceous Ore + 1 Refined Metal", "2 Carbon Mesh", 2, "Carbon Mesh", "Signature: +30% near Carbon Mine"),
    ("T2", "Pressure-Valve Forge", 120, "2 Refined Metal + 1 Nickel Ore", "1 Pressure Valves", 1, "Pressure Valves", ""),
    ("T2", "Textile Mill", 100, "2 Carbon Mesh + 1 Hydroponic Yield", "2 Textiles", 2, "Textiles", ""),
    ("T2", "Furnishings Workshop", 150, "1 Aluminum + 1 Carbon Mesh + 1 Textiles", "1 Furnishings", 1, "Furnishings", ""),
    ("T2", "Distillery", 180, "3 Hydroponic Yield + 1 Water Ice", "1 Spirits", 1, "Spirits", ""),
    ("T3", "Trace Mine", 90, "—", "2 Rare Trace Elements", 2, "Rare Trace Elements", "Rare-trace NEAs only"),
    ("T3", "Logic Foundry", 120, "1 Refined Metal + 2 Rare Trace Elements + 1 Aluminum", "2 Logic Boards", 2, "Logic Boards", "Signature: +30% near Trace Mine"),
    ("T3", "Comms Workshop", 100, "2 Logic Boards + 1 Aluminum", "1 Comms Module", 1, "Comms Module", ""),
    ("T3", "Sensor Lab", 110, "2 Logic Boards + 1 Carbon Mesh", "1 Sensor Array", 1, "Sensor Array", ""),
    ("T3", "Maintenance Shop", 90, "1 Pressure Valves + 1 Aluminum", "1 Maintenance Kit", 1, "Maintenance Kit", ""),
    ("T3", "Catalytic Reactor", 120, "2 Ammonia + 1 Hydrogen Fuel", "2 Refined Hydrogen", 2, "Refined Hydrogen", ""),
    ("T3", "Automation Hub", 0, "—", "hosts 4 automation rules", 0, "", "Required for any automation at body"),
]

# Build a Resources lookup map (resource name -> Resources!G row for sell price)
res_row_map = {r[1]: i + 2 for i, r in enumerate(resources)}

for i, (tier, name, cycle, inputs, outputs, out_qty, out_res, notes) in enumerate(recipes, start=2):
    ws.cell(row=i, column=1, value=tier).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=name).font = Font(name=FONT)
    ws.cell(row=i, column=3, value=cycle if cycle else None).font = INPUT_BLUE if cycle else Font(name=FONT)
    ws.cell(row=i, column=4, value=inputs).font = Font(name=FONT)
    ws.cell(row=i, column=5, value=outputs).font = Font(name=FONT)
    ws.cell(row=i, column=6, value=out_qty if out_qty else None).font = INPUT_BLUE if out_qty else Font(name=FONT)
    ws.cell(row=i, column=7, value=out_res).font = Font(name=FONT)
    if cycle and out_qty:
        ws.cell(row=i, column=8, value=f"=60/C{i}").font = FORMULA_BLACK
        ws.cell(row=i, column=8).number_format = "0.00"
        ws.cell(row=i, column=9, value=f"=F{i}*60/C{i}").font = FORMULA_BLACK
        ws.cell(row=i, column=9).number_format = "0.00"
        if out_res in res_row_map:
            res_row = res_row_map[out_res]
            ws.cell(row=i, column=10, value=f"=Resources!G{res_row}").font = LINK_GREEN
            ws.cell(row=i, column=10).number_format = "$#,##0;($#,##0);-"
            ws.cell(row=i, column=11, value=f"=I{i}*J{i}").font = FORMULA_BLACK
            ws.cell(row=i, column=11).number_format = "$#,##0;($#,##0);-"
        else:
            ws.cell(row=i, column=10, value="—").font = Font(name=FONT)
            ws.cell(row=i, column=11, value="—").font = Font(name=FONT)
    else:
        for j in (8, 9, 10, 11):
            ws.cell(row=i, column=j, value="—").font = Font(name=FONT)
    ws.cell(row=i, column=12, value=notes).font = Font(name=FONT)
    for j in range(1, 13):
        ws.cell(row=i, column=j).border = BORDER

footer = len(recipes) + 3
ws.cell(row=footer, column=1, value="Cycles/min = 60 / cycle_s. Output rate/min = output_qty × 60 / cycle_s.").font = NOTE
ws.cell(row=footer + 1, column=1, value="Output value/min uses Earth Sell price as a unit reference — the actual route economics depend on hull, leg, and fuel costs (see Fuel_Costs / Ships).").font = NOTE
ws.cell(row=footer + 2, column=1, value="Recipes for Electrolyzer also produce 1 Oxygen byproduct per cycle — not modeled in the headline rate.").font = NOTE

# =============================================================================
# 5. ADJACENCY PAIRS (R72)
# =============================================================================
ws = wb.create_sheet("Adjacency_Pairs")
header_row(ws, 1, ["Tier Band", "Placer", "Pairs With", "Bonus", "Notes"], [12, 28, 28, 9, 50])

pairs = [
    ("T0", "Smelter", "Small Mine", 0.30, "Signature T0 pair; FTUE compare-moment"),
    ("T1", "Smelter", "Lunar Surface Mine", 0.25, "T1 variant"),
    ("T0", "Electrolyzer", "Ice Mine", 0.25, "Signature T0 fluid pair"),
    ("T1", "Refinery (Aluminum)", "Lunar Surface Mine", 0.30, "Signature T1 pair"),
    ("T1", "Construction Yard", "Refinery (Aluminum)", 0.20, "Mid-chain feed"),
    ("T1", "Construction Yard", "Smelter", 0.20, "Alt feed source"),
    ("T1", "Habitat Assembler", "Construction Yard", 0.25, "Late-T1 chain capper"),
    ("T1", "Greenhouse (small)", "Ice Mine", 0.20, "Cuts water-feed friction"),
    ("T2", "Hydroponics Bay", "Ice Mine", 0.20, "T2 upgrade path"),
    ("T2", "Hydroponic Greenhouse", "Hydroponics Bay", 0.25, "T2 stack"),
    ("T2", "Glass Furnace", "NEA Mine (Silicates)", 0.30, "Signature T2 pair"),
    ("T2", "Carbon Mill", "NEA Mine (Carbon)", 0.30, "Signature T2 pair"),
    ("T2", "Pressure-Valve Forge", "Smelter", 0.20, "Multi-input chain helper"),
    ("T2", "Textile Mill", "Hydroponic Greenhouse", 0.20, "Comfort-tier feed"),
    ("T2", "Furnishings Workshop", "Carbon Mill", 0.20, "Multi-feed chain capper"),
    ("T2", "Distillery", "Hydroponic Greenhouse", 0.25, "Endgame-T2 luxury chain"),
    ("T3", "Logic Foundry", "Trace Mine", 0.30, "Signature T3 pair"),
    ("T3", "Comms Workshop", "Logic Foundry", 0.25, "Mid-T3 chain stack"),
    ("T3", "Sensor Lab", "Logic Foundry", 0.25, "Mid-T3 chain stack"),
    ("T3", "Maintenance Shop", "Pressure-Valve Forge", 0.20, "Cross-tier (T2/T3)"),
    ("T3", "Catalytic Reactor", "Electrolyzer", 0.20, "Fluid-stack pair"),
    ("T3", "Automation Hub", "(any building)", 0.10, "On-site staff coordination; deliberately small"),
]
for i, (tier, placer, pair, bonus, notes) in enumerate(pairs, start=2):
    ws.cell(row=i, column=1, value=tier).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=placer).font = Font(name=FONT)
    ws.cell(row=i, column=3, value=pair).font = Font(name=FONT)
    ws.cell(row=i, column=4, value=bonus).font = INPUT_BLUE
    ws.cell(row=i, column=4).number_format = "0%"
    ws.cell(row=i, column=5, value=notes).font = Font(name=FONT)
    for j in range(1, 6):
        ws.cell(row=i, column=j).border = BORDER
footer = len(pairs) + 3
ws.cell(row=footer, column=1, value="Range: 15%–35% (R72). Signature pairs sit at top (30%); standard pairs at 20–25%.").font = NOTE
ws.cell(row=footer + 1, column=1, value="Stacking cap: max 2 pair bonuses per building. Pairs are bidirectional (each side receives the listed bonus).").font = NOTE
ws.cell(row=footer + 2, column=1, value="Storage buildings (Silo / Tank / Cryo Tank) are neutral — don't grant or receive bonuses (R70).").font = NOTE
ws.cell(row=footer + 3, column=1, value="Collaboration radius: default 2 tiles, uniform (R69). Per-building override architecturally supported.").font = NOTE

# =============================================================================
# 6. PEOPLE COSTS (R73)
# =============================================================================
ws = wb.create_sheet("People_Costs")
header_row(ws, 1, ["Tier Band", "Building", "People Cost"], [12, 30, 14])
people = [
    ("Source", "Earth-orbit starter crew (free, T0)", -8),  # negative = capacity, marker
    ("T0", "Small Mine", 2),
    ("T0", "Ice Mine", 2),
    ("T0", "Smelter", 3),
    ("T0", "Electrolyzer", 3),
    ("T0", "Probe Bay", 1),
    ("T1", "Lunar Surface Mine", 3),
    ("T1", "Refinery (Aluminum)", 4),
    ("T1", "Construction Yard", 5),
    ("T1", "Habitat Assembler", 6),
    ("T1", "Greenhouse (small)", 2),
    ("T1", "Life Support (Water / Oxygen / Food, each)", 1),
    ("T2", "NEA Mine (any)", 3),
    ("T2", "Hydroponics Bay", 3),
    ("T2", "Hydroponic Greenhouse", 3),
    ("T2", "Glass Furnace", 5),
    ("T2", "Carbon Mill", 4),
    ("T2", "Pressure-Valve Forge", 5),
    ("T2", "Textile Mill", 4),
    ("T2", "Furnishings Workshop", 6),
    ("T2", "Distillery", 4),
    ("T3", "Trace Mine", 3),
    ("T3", "Logic Foundry", 5),
    ("T3", "Comms Workshop", 5),
    ("T3", "Sensor Lab", 5),
    ("T3", "Maintenance Shop", 4),
    ("T3", "Catalytic Reactor", 5),
    ("T3", "Automation Hub", 1),
    ("Storage", "Silo / Tank / Cryo Tank", 0),
]
for i, (tier, name, cost) in enumerate(people, start=2):
    ws.cell(row=i, column=1, value=tier).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=name).font = Font(name=FONT)
    if cost < 0:
        ws.cell(row=i, column=3, value=abs(cost)).font = INPUT_BLUE
        ws.cell(row=i, column=3).number_format = '"+"0" cap"'
    else:
        ws.cell(row=i, column=3, value=cost).font = INPUT_BLUE
    for j in range(1, 4):
        ws.cell(row=i, column=j).border = BORDER

footer = len(people) + 3
ws.cell(row=footer, column=1, value="People Capacity sources & multipliers").font = SECTION
ws.cell(row=footer + 1, column=1, value="Source").font = HEADER
ws.cell(row=footer + 1, column=1).fill = HEADER_FILL
ws.cell(row=footer + 1, column=2, value="Contribution").font = HEADER
ws.cell(row=footer + 1, column=2).fill = HEADER_FILL
ws.cell(row=footer + 2, column=1, value="Earth-orbit starter").font = Font(name=FONT)
ws.cell(row=footer + 2, column=2, value="8 capacity (free, T0)").font = Font(name=FONT)
ws.cell(row=footer + 3, column=1, value="Each colony").font = Font(name=FONT)
ws.cell(row=footer + 3, column=2, value="Pop × pop-tier multiplier (see Pop_Tiers)").font = Font(name=FONT)

# Sanity check — T0 starter chain
sc = footer + 5
ws.cell(row=sc, column=1, value="T0 sanity check").font = SECTION
ws.cell(row=sc + 1, column=1, value="Available capacity").font = Font(name=FONT)
ws.cell(row=sc + 1, column=3, value=8).font = INPUT_BLUE
ws.cell(row=sc + 2, column=1, value="1 Small Mine + 1 Smelter + 1 Probe Bay").font = Font(name=FONT)
ws.cell(row=sc + 2, column=3, value=f"=C3+C5+C7").font = FORMULA_BLACK
ws.cell(row=sc + 3, column=1, value="Reserve").font = Font(name=FONT)
ws.cell(row=sc + 3, column=3, value=f"=C{sc + 1}-C{sc + 2}").font = FORMULA_BLACK

# =============================================================================
# 7. POP TIERS
# =============================================================================
ws = wb.create_sheet("Pop_Tiers")
header_row(ws, 1, ["Pop Tier", "Unlocks At", "Multiplier", "Settle-in Window", "Continuous Needs",
                   "Growth-Tier Bundle (one-time)", "Sample 32-pop capacity"],
           [14, 14, 12, 18, 40, 40, 18])
pop_tiers = [
    ("Survival", "T1 (first habitat)", 1.00, "5 min", "Water, Oxygen, Food Pack", "4 Construction Materials"),
    ("Settled", "T1 (habitat upgraded)", 1.25, "20 min", "(no new continuous needs)", "8 Construction Materials + 2 Habitat Module"),
    ("Growing", "T2 (Glass + Valves available)", 1.50, "1 h", "+ Pressure Valves drip (1 / pop / 30 min)", "6 Habitat Glass + 4 Pressure Valves"),
    ("Comfortable", "T2 late", 1.75, "2 h", "+ Textiles drip (1 / pop / 60 min)", "8 Textiles + 4 Furnishings"),
    ("Affluent", "T2 endgame", 2.00, "4 h", "+ Spirits drip (1 / pop / 90 min)", "6 Furnishings + 4 Spirits"),
    ("Networked (T3)", "T3 endgame", 2.25, "8 h", "+ Comms Module drip (1 / pop / 60 min) + Sensor Array drip (1 / pop / 240 min)", "6 Comms Module + 4 Sensor Array"),
]
for i, (tier, unlock, mult, window, needs, bundle) in enumerate(pop_tiers, start=2):
    ws.cell(row=i, column=1, value=tier).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=unlock).font = Font(name=FONT)
    ws.cell(row=i, column=3, value=mult).font = INPUT_BLUE
    ws.cell(row=i, column=3).number_format = '0.00"×"'
    ws.cell(row=i, column=4, value=window).font = Font(name=FONT)
    ws.cell(row=i, column=5, value=needs).font = Font(name=FONT)
    ws.cell(row=i, column=6, value=bundle).font = Font(name=FONT)
    ws.cell(row=i, column=7, value=f"=32*C{i}").font = FORMULA_BLACK
    ws.cell(row=i, column=7).number_format = "0.0"
    for j in range(1, 8):
        ws.cell(row=i, column=j).border = BORDER

footer = len(pop_tiers) + 3
ws.cell(row=footer, column=1, value="Settle-in across AFK: progresses while away, capped at one tier transition per AFK return (R67). Window resumes (does not restart).").font = NOTE
ws.cell(row=footer + 1, column=1, value="Sample 32-pop capacity column: a hypothetical 32-population habitat at this tier — convenient ratio check vs. building people-cost totals.").font = NOTE

# =============================================================================
# 8. STORAGE CAPS
# =============================================================================
ws = wb.create_sheet("Storage_Caps")
header_row(ws, 1, ["Storage Building", "Class", "T0 cap", "T2 cap", "T4 cap", "T6 cap", "T2/T0 multiplier"],
           [22, 16, 10, 10, 10, 10, 18])
storage = [
    ("Silo", "Solids", 300, 900, 3000, 9000),
    ("Tank", "Fluids/Gases", 180, 540, 1800, 5400),
    ("Cryo Tank", "Specialty cold-chain", None, 240, 800, 2400),
]
for i, (name, cls, t0, t2, t4, t6) in enumerate(storage, start=2):
    ws.cell(row=i, column=1, value=name).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=cls).font = Font(name=FONT)
    ws.cell(row=i, column=3, value=t0).font = INPUT_BLUE if t0 else Font(name=FONT)
    if not t0:
        ws.cell(row=i, column=3, value="—").font = Font(name=FONT)
    ws.cell(row=i, column=4, value=t2).font = INPUT_BLUE
    ws.cell(row=i, column=5, value=t4).font = INPUT_BLUE
    ws.cell(row=i, column=6, value=t6).font = INPUT_BLUE
    if t0:
        ws.cell(row=i, column=7, value=f"=D{i}/C{i}").font = FORMULA_BLACK
        ws.cell(row=i, column=7).number_format = '0.0"×"'
    else:
        ws.cell(row=i, column=7, value="—").font = Font(name=FONT)
    for j in range(1, 8):
        ws.cell(row=i, column=j).border = BORDER

footer = len(storage) + 3
ws.cell(row=footer, column=1, value="Each storage building takes 1 grid slot. Capacity is per-resource (each Silo holds a single resource at a time, R74).").font = NOTE
ws.cell(row=footer + 1, column=1, value="Auto-upgrade in place at tier transitions (R60) — no demolish/rebuild, no resource loss. Reassign empty storage for $200 (placeholder).").font = NOTE

# =============================================================================
# 9. FUEL COSTS (R74)
# =============================================================================
ws = wb.create_sheet("Fuel_Costs")
header_row(ws, 1, ["Leg", "Base Fuel"], [40, 14])
fuel_legs = [
    ("Earth ↔ Moon", 3),
    ("Earth ↔ NEA cluster", 5),
    ("NEA ↔ Lunar Habitat", 2),
    ("Earth ↔ Mars (window-good)", 25),
    ("Earth ↔ Mars (window-poor)", 40),
    ("Earth ↔ Belt", 60),
    ("Earth ↔ Jovian", 140),
]
for i, (leg, fuel) in enumerate(fuel_legs, start=2):
    ws.cell(row=i, column=1, value=leg).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=fuel).font = INPUT_BLUE
    for j in range(1, 3):
        ws.cell(row=i, column=j).border = BORDER

# Window multipliers
fr = len(fuel_legs) + 3
ws.cell(row=fr, column=1, value="Window multipliers").font = SECTION
ws.cell(row=fr + 1, column=1, value="Good window").font = Font(name=FONT)
ws.cell(row=fr + 1, column=2, value=0.85).font = INPUT_BLUE
ws.cell(row=fr + 1, column=2).number_format = "0.00"
ws.cell(row=fr + 2, column=1, value="Neutral").font = Font(name=FONT)
ws.cell(row=fr + 2, column=2, value=1.00).font = INPUT_BLUE
ws.cell(row=fr + 2, column=2).number_format = "0.00"
ws.cell(row=fr + 3, column=1, value="Poor window").font = Font(name=FONT)
ws.cell(row=fr + 3, column=2, value=1.25).font = INPUT_BLUE
ws.cell(row=fr + 3, column=2).number_format = "0.00"

# Sample computation: total fuel for a Hauler-1 (1.00× hull mult) on each leg, neutral
sr = fr + 5
ws.cell(row=sr, column=1, value="Sample: Hauler-1 fuel cost per leg (neutral window)").font = SECTION
header_row(ws, sr + 1, ["Leg", "Base", "Hull mult (Hauler-1)", "Window mult", "Total fuel"], None)
hauler_mult = 1.00
for i, (leg, fuel) in enumerate(fuel_legs, start=sr + 2):
    ws.cell(row=i, column=1, value=leg).font = Font(name=FONT)
    base_row = i - sr - 1 + 1  # base fuel row in the original section
    ws.cell(row=i, column=2, value=f"=B{base_row + 1}").font = LINK_GREEN
    ws.cell(row=i, column=3, value=hauler_mult).font = INPUT_BLUE
    ws.cell(row=i, column=3).number_format = '0.00"×"'
    ws.cell(row=i, column=4, value=f"=B{fr + 2}").font = LINK_GREEN
    ws.cell(row=i, column=4).number_format = '0.00"×"'
    ws.cell(row=i, column=5, value=f"=B{i}*C{i}*D{i}").font = FORMULA_BLACK
    ws.cell(row=i, column=5).number_format = "0.0"
    for j in range(1, 6):
        ws.cell(row=i, column=j).border = BORDER

footer = sr + 2 + len(fuel_legs) + 1
ws.cell(row=footer, column=1, value="Fuel deducts from origin body's tank at dispatch. Multi-stop: per-leg deduction at each leg's origin.").font = NOTE
ws.cell(row=footer + 1, column=1, value="Mid-route empty origin → ship strands; recovery: dispatch tanker (or buy Earth-bought Hydrogen Fuel order).").font = NOTE

# =============================================================================
# 10. SHIPS
# =============================================================================
ws = wb.create_sheet("Ships")
header_row(ws, 1, ["Hull", "Tier", "Family", "Solid slots", "Fluid slots", "Total slots",
                   "Speed", "Fuel/Route", "Earth Buy", "$ per slot"], [12, 6, 22, 12, 12, 12, 8, 12, 12, 12])
ships = [
    ("Hauler-1", "T0", "Specialized Solid", 30, 0, 1.00, 1.00, 3000),
    ("Mixer-1", "T0", "Combined", 20, 10, 0.95, 1.05, 4200),
    ("Tanker-1", "T1", "Specialized Fluid", 0, 25, 0.90, 1.10, 4800),
    ("Hauler-2", "T2", "Specialized Solid", 75, 0, 1.10, 1.30, 9500),
    ("Tanker-2", "T2", "Specialized Fluid", 0, 60, 1.00, 1.40, 11500),
    ("Mixer-2", "T2", "Combined", 45, 25, 1.00, 1.40, 12500),
]
for i, (name, tier, fam, solid, fluid, speed, fuel, cost) in enumerate(ships, start=2):
    ws.cell(row=i, column=1, value=name).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=tier).font = Font(name=FONT)
    ws.cell(row=i, column=3, value=fam).font = Font(name=FONT)
    ws.cell(row=i, column=4, value=solid).font = INPUT_BLUE
    ws.cell(row=i, column=5, value=fluid).font = INPUT_BLUE
    ws.cell(row=i, column=6, value=f"=D{i}+E{i}").font = FORMULA_BLACK
    ws.cell(row=i, column=7, value=speed).font = INPUT_BLUE
    ws.cell(row=i, column=7).number_format = '0.00"×"'
    ws.cell(row=i, column=8, value=fuel).font = INPUT_BLUE
    ws.cell(row=i, column=8).number_format = '0.00"×"'
    ws.cell(row=i, column=9, value=cost).font = INPUT_BLUE
    ws.cell(row=i, column=9).number_format = "$#,##0;($#,##0);-"
    ws.cell(row=i, column=10, value=f"=I{i}/F{i}").font = FORMULA_BLACK
    ws.cell(row=i, column=10).number_format = "$#,##0.00;($#,##0.00);-"
    for j in range(1, 11):
        ws.cell(row=i, column=j).border = BORDER

footer = len(ships) + 3
ws.cell(row=footer, column=1, value="$ per slot = Earth Buy ÷ Total slots — efficiency-vs-investment metric for fleet purchase decisions.").font = NOTE
ws.cell(row=footer + 1, column=1, value="Specialist hulls (probe ships, builders) deferred — at T0–T2 the Probe Bay building handles surveys.").font = NOTE
ws.cell(row=footer + 2, column=1, value="T2 hulls roughly 3× capacity at ~3× cost — meaningful investment, not free upgrade.").font = NOTE

# =============================================================================
# 11. TIER GATES (R76)
# =============================================================================
ws = wb.create_sheet("Tier_Gates")
header_row(ws, 1, ["Transition", "Clause", "Theme", "Primary Path", "Alternate Path"],
           [12, 8, 22, 50, 50])
gates = [
    ("T0 → T1 (Lunar Foothold)", "A", "Refined output", "Sell 200 Refined Metal to Earth", "Sell 800 Iron Ore raw to Earth"),
    ("T0 → T1 (Lunar Foothold)", "B", "Fuel readiness", "Accumulate 50 Hydrogen Fuel reserves at any single body", "Sell 80 Hydrogen Fuel cumulatively to Earth"),
    ("T1 → T2 (NEA Industry)", "A", "Settlement", "First habitat reaches Pop 50", "Maintain Survival tier on 2 distinct habitats simultaneously"),
    ("T1 → T2 (NEA Industry)", "B", "Frontier reach", "Claim 2 NEA surveys", "Claim 1 NEA + 1 lunar surface site"),
    ("T2 → T3 (Cislunar Network)", "A", "Local life support", "Local oxygen break-even at lunar habitat (no Earth O2 imports for 24h game time)", "Ship 200 locally-produced Oxygen to a different body cumulatively"),
    ("T2 → T3 (Cislunar Network)", "B", "Pop maturity", "Habitat reaches Comfortable pop tier", "Growing pop tier on 2 habitats simultaneously"),
    ("T3 → T4 (Martian Reach)", "A", "Distributed production", "Build first non-Earth shipyard at Mars orbit", "Establish Industrial Control Unit production at 2 distinct bodies (T4 placeholder)"),
    ("T3 → T4 (Martian Reach)", "B", "Networked colony", "Maintain a Networked-tier colony", "Maintain Affluent + 2 Comfortable colonies simultaneously"),
]
for i, (transition, clause, theme, primary, alt) in enumerate(gates, start=2):
    ws.cell(row=i, column=1, value=transition).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=clause).font = Font(name=FONT)
    ws.cell(row=i, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=i, column=3, value=theme).font = Font(name=FONT)
    ws.cell(row=i, column=4, value=primary).font = Font(name=FONT)
    ws.cell(row=i, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=i, column=5, value=alt).font = Font(name=FONT)
    ws.cell(row=i, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    for j in range(1, 6):
        ws.cell(row=i, column=j).border = BORDER
    ws.row_dimensions[i].height = 36
footer = len(gates) + 3
ws.cell(row=footer, column=1, value="Both clauses must complete (primary OR alternate within each clause). Alternates are equivalent-difficulty-different-approach (R76).").font = NOTE
ws.cell(row=footer + 1, column=1, value="Stage 3 acceptance: 10–25% of runs reach the alternate path (signal it's a real lane, not a curiosity).").font = NOTE
ws.cell(row=footer + 2, column=1, value="T4+ gates remain at named-only level (Pending #39).").font = NOTE

# =============================================================================
# 12. ACTION DURATIONS
# =============================================================================
ws = wb.create_sheet("Action_Durations")
header_row(ws, 1, ["Action", "Duration", "Seconds (parsed)"], [40, 16, 16])
actions = [
    ("Survey one NEA (T0 probe)", "4 min", 240),
    ("Mine cycle (small mine, 1 batch)", "30 s", 30),
    ("Smelt cycle", "45 s", 45),
    ("Build small habitat", "8 min", 480),
    ("Earth → Moon transit (good window)", "6 min", 360),
    ("Earth → Mars transit (good window)", "35 min", 2100),
    ("Earth → Belt transit", "90 min", 5400),
    ("Build T1 ship at Earth shipyard", "12 min", 720),
    ("Pop-tier settle-in: Survival", "5 min", 300),
    ("Pop-tier settle-in: Settled", "20 min", 1200),
    ("Pop-tier settle-in: Growing", "1 h", 3600),
    ("Pop-tier settle-in: Comfortable", "2 h", 7200),
    ("Pop-tier settle-in: Affluent", "4 h", 14400),
    ("Pop-tier settle-in: Networked", "8 h", 28800),
    ("AFK catch-up hard cap", "24 h", 86400),
]
for i, (action, dur, secs) in enumerate(actions, start=2):
    ws.cell(row=i, column=1, value=action).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=dur).font = Font(name=FONT)
    ws.cell(row=i, column=3, value=secs).font = INPUT_BLUE
    for j in range(1, 4):
        ws.cell(row=i, column=j).border = BORDER

footer = len(actions) + 3
ws.cell(row=footer, column=1, value="Game time : real time = 1:1. Simulation tick = 1 Hz foreground; deterministic catch-up offline.").font = NOTE

# =============================================================================
# 13. EARTH PREFAB KITS
# =============================================================================
ws = wb.create_sheet("Earth_Prefab_Kits")
header_row(ws, 1, ["Tier", "Kit Name", "Cost", "Contents", "Notes"], [6, 32, 12, 50, 30])
kits = [
    ("T1", "Lunar Habitat Kit", None, "1 Habitat Module (drops directly on Moon)", "Bootstraps first habitat before local construction chain"),
    ("T1", "Lunar Surface Mine Kit", None, "1 Lunar Surface Mine (drops on Moon)", "Bootstraps first local mine"),
    ("T3", "Cislunar Comms Relay Kit", 28000, "1 Comms Workshop + 1 Automation Hub", "Bootstraps automation on first non-lunar T3 expansion"),
    ("T3", "Sensor Network Kit", 22000, "1 Sensor Lab + 2 Maintenance Shops", "Bootstraps fleet maintenance scaling"),
    ("T4", "Mars Foothold Kit", None, "(deferred to T4 drill)", "Pending #39"),
]
for i, (tier, name, cost, contents, notes) in enumerate(kits, start=2):
    ws.cell(row=i, column=1, value=tier).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=name).font = Font(name=FONT)
    if cost:
        ws.cell(row=i, column=3, value=cost).font = INPUT_BLUE
        ws.cell(row=i, column=3).number_format = "$#,##0;($#,##0);-"
    else:
        ws.cell(row=i, column=3, value="(unspec)").font = Font(name=FONT)
    ws.cell(row=i, column=4, value=contents).font = Font(name=FONT)
    ws.cell(row=i, column=5, value=notes).font = Font(name=FONT)
    for j in range(1, 6):
        ws.cell(row=i, column=j).border = BORDER

footer = len(kits) + 3
ws.cell(row=footer, column=1, value="Hand-authored, 1-of-1 per kit per tier (R37). Each kit = a tier-up payoff moment.").font = NOTE
ws.cell(row=footer + 1, column=1, value="Total catalog target: ~10–14 kits across T1–T6 (Content Targets).").font = NOTE

# =============================================================================
# 14. T1_T2_CHAIN_MATH — 60-pop Comfortable colony build-out drill
# =============================================================================
ws = wb.create_sheet("T1_T2_Chain_Math")
SECTION_13 = Font(name=FONT, bold=True, size=13)


def cm_section(sheet, row, label, span=8):
    c = sheet.cell(row=row, column=1, value=label)
    c.font = SECTION_13
    c.fill = SECTION_FILL
    for j in range(2, span + 1):
        sheet.cell(row=row, column=j).fill = SECTION_FILL


widths = [32, 16, 14, 14, 14, 14, 14, 28]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws["A1"] = "T1–T2 Math Drill: 60-pop Comfortable Lunar Habitat"
ws["A1"].font = Font(name=FONT, bold=True, size=16)
ws["A2"] = "Target: validate that placeholder numbers can sustain a Comfortable colony on a 7×7 lunar grid (~25–49 slots) within People Capacity at this tier."
ws["A2"].font = NOTE
ws["A3"] = "Edit blue inputs to test sensitivity. Black formulas update; the slack readouts at the bottom flag bottlenecks."
ws["A3"].font = NOTE

cm_section(ws, 5, "1. Scenario Inputs")
ws["A6"] = "Population"
ws["B6"] = 60
ws["B6"].font = INPUT_BLUE
ws["A7"] = "Pop tier"
ws["B7"] = "Comfortable"
ws["A8"] = "People Capacity multiplier"
ws["B8"] = 1.75
ws["B8"].font = INPUT_BLUE
ws["B8"].number_format = '0.00"×"'
ws["A9"] = "People Capacity available"
ws["B9"] = "=B6*B8"
ws["B9"].font = FORMULA_BLACK
ws["B9"].number_format = "0.0"
ws["C9"] = "(Pop × multiplier; Earth-orbit starter +8 not included — single-colony view)"
ws["C9"].font = NOTE
ws["A10"] = "Grid slots (assume 7×7 lunar habitat, mid-roll)"
ws["B10"] = 49
ws["B10"].font = INPUT_BLUE
ws["A11"] = "Reserved for habitat life support buildings (Water/O2/Food)"
ws["B11"] = 3
ws["B11"].font = INPUT_BLUE
ws["A12"] = "Reserved for storage (Silos + Tank)"
ws["B12"] = 4
ws["B12"].font = INPUT_BLUE
ws["A13"] = "Slots available for production"
ws["B13"] = "=B10-B11-B12"
ws["B13"].font = FORMULA_BLACK

cm_section(ws, 15, "2. Continuous Per-Pop Demand → Total Rate Needed")
header_row(ws, 16, ["Need", "Spec (per pop)", "Per-pop /min", "Total /min @ 60 pop", "", "", "", "Notes"])

needs = [
    ("Water Ice", "1 / pop / 8 min", 1 / 8, "Life Support — Water"),
    ("Oxygen", "1 / pop / 6 min", 1 / 6, "Life Support — Oxygen"),
    ("Food Pack", "1 / pop / 12 min", 1 / 12, "Life Support — Food"),
    ("Pressure Valves", "1 / pop / 30 min (Growing tier+)", 1 / 30, "Drip"),
    ("Textiles", "1 / pop / 60 min (Comfortable tier+)", 1 / 60, "Drip"),
]
for i, (need, spec, rate_pp, notes) in enumerate(needs, start=17):
    ws.cell(row=i, column=1, value=need).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=spec).font = Font(name=FONT)
    ws.cell(row=i, column=3, value=rate_pp).font = INPUT_BLUE
    ws.cell(row=i, column=3).number_format = "0.0000"
    ws.cell(row=i, column=4, value=f"=$B$6*C{i}").font = FORMULA_BLACK
    ws.cell(row=i, column=4).number_format = "0.00"
    ws.cell(row=i, column=8, value=notes).font = NOTE
    for j in range(1, 9):
        ws.cell(row=i, column=j).border = BORDER

cm_section(ws, 24, "3. Building Throughput → Buildings Required (ceiling)")
header_row(ws, 25, ["Building", "Cycle (s)", "Output qty/cycle", "Output rate/min", "Demand /min",
                    "Buildings needed", "People per", "Total people"])

buildings = [
    ("Ice Mine", 40, 8, 17, 2),
    ("Electrolyzer (for Oxygen)", 60, 1, 18, 3),
    ("Hydroponic Greenhouse (T2 upgrade)", 45, 4, 19, 3),
    ("Hydroponics Bay (input to Hydro Greenhouse)", 60, 2, None, 3),
    ("Pressure-Valve Forge", 120, 1, 20, 5),
    ("Textile Mill", 100, 2, 21, 4),
]
for i, (name, cycle, out_qty, dem_row, people) in enumerate(buildings, start=26):
    ws.cell(row=i, column=1, value=name).font = Font(name=FONT)
    ws.cell(row=i, column=2, value=cycle).font = INPUT_BLUE
    ws.cell(row=i, column=3, value=out_qty).font = INPUT_BLUE
    ws.cell(row=i, column=4, value=f"=C{i}*60/B{i}").font = FORMULA_BLACK
    ws.cell(row=i, column=4).number_format = "0.00"
    if dem_row is not None:
        ws.cell(row=i, column=5, value=f"=D{dem_row}").font = LINK_GREEN
        ws.cell(row=i, column=5).number_format = "0.00"
        ws.cell(row=i, column=6, value=f"=ROUNDUP(E{i}/D{i},0)").font = FORMULA_BLACK
    else:
        ws.cell(row=i, column=5, value=f"=2*60/B28*F28").font = FORMULA_BLACK
        ws.cell(row=i, column=5).number_format = "0.00"
        ws.cell(row=i, column=6, value=f"=ROUNDUP(E{i}/D{i},0)").font = FORMULA_BLACK
    ws.cell(row=i, column=7, value=people).font = INPUT_BLUE
    ws.cell(row=i, column=8, value=f"=F{i}*G{i}").font = FORMULA_BLACK
    for j in range(1, 9):
        ws.cell(row=i, column=j).border = BORDER

cm_section(ws, 33, "4. Upstream feed required (raw + intermediate)")
header_row(ws, 34, ["Building", "Cycle (s)", "Output qty/cycle", "Output rate/min", "Demand /min",
                    "Buildings needed", "People per", "Total people"])

upstream = [
    ("Refined Metal needed (for PV Forges)", "—", "—", "=2*60/B30*F30", None),
    ("Smelter (5 Iron Ore → 2 Refined Metal)", 45, 2, None, 3),
    ("Iron Ore needed (for Smelters)", "—", "—", None, None),
    ("Small Mine (10 Iron Ore)", 30, 10, None, 2),
    ("Nickel Ore needed (for PV Forges)", "—", "—", "=1*60/B30*F30", None),
    ("NEA Mine (Nickel)", 45, 6, None, 3),
    ("Carbon Mesh needed (for Textile Mills)", "—", "—", "=2*60/B31*F31", None),
    ("Carbon Mill (3 Carbon Ore + 1 RM → 2 Carbon Mesh)", 75, 2, None, 4),
    ("Carbonaceous Ore needed (for Carbon Mill)", "—", "—", None, None),
    ("NEA Mine (Carbon)", 45, 6, None, 3),
]

for i, (name, cycle, out_qty, demand, people) in enumerate(upstream, start=35):
    ws.cell(row=i, column=1, value=name).font = Font(name=FONT)
    if cycle != "—":
        ws.cell(row=i, column=2, value=cycle).font = INPUT_BLUE
        ws.cell(row=i, column=3, value=out_qty).font = INPUT_BLUE
        ws.cell(row=i, column=4, value=f"=C{i}*60/B{i}").font = FORMULA_BLACK
        ws.cell(row=i, column=4).number_format = "0.00"
    else:
        ws.cell(row=i, column=2, value="—").font = Font(name=FONT)
        ws.cell(row=i, column=3, value="—").font = Font(name=FONT)
        ws.cell(row=i, column=4, value="—").font = Font(name=FONT)
    if demand:
        ws.cell(row=i, column=5, value=demand).font = FORMULA_BLACK
        ws.cell(row=i, column=5).number_format = "0.00"
    if people is not None:
        ws.cell(row=i, column=7, value=people).font = INPUT_BLUE
    for j in range(1, 9):
        ws.cell(row=i, column=j).border = BORDER

# Wire dependencies
ws["E36"] = "=E35"
ws["E36"].font = LINK_GREEN
ws["E36"].number_format = "0.00"
ws["F36"] = "=ROUNDUP(E36/D36,0)"
ws["F36"].font = FORMULA_BLACK
ws["H36"] = "=F36*G36"
ws["H36"].font = FORMULA_BLACK

ws["E37"] = "=5*60/B36*F36"
ws["E37"].font = FORMULA_BLACK
ws["E37"].number_format = "0.00"
ws["F38"] = "=ROUNDUP(E37/D38,0)"
ws["F38"].font = FORMULA_BLACK
ws["H38"] = "=F38*G38"
ws["H38"].font = FORMULA_BLACK

ws["F40"] = "=ROUNDUP(E39/D40,0)"
ws["F40"].font = FORMULA_BLACK
ws["H40"] = "=F40*G40"
ws["H40"].font = FORMULA_BLACK

ws["E42"] = "=E41"
ws["E42"].font = LINK_GREEN
ws["E42"].number_format = "0.00"
ws["F42"] = "=ROUNDUP(E42/D42,0)"
ws["F42"].font = FORMULA_BLACK
ws["H42"] = "=F42*G42"
ws["H42"].font = FORMULA_BLACK

ws["E43"] = "=3*60/B42*F42"
ws["E43"].font = FORMULA_BLACK
ws["E43"].number_format = "0.00"
ws["F44"] = "=ROUNDUP(E43/D44,0)"
ws["F44"].font = FORMULA_BLACK
ws["H44"] = "=F44*G44"
ws["H44"].font = FORMULA_BLACK

cm_section(ws, 46, "5. Roll-up: people-cost and slot-count totals")
header_row(ws, 47, ["Category", "Buildings (count)", "People used", "Slots used", "", "", "", ""])

chain_rows_3 = [26, 27, 28, 29, 30, 31]
chain_rows_4 = [36, 38, 40, 42, 44]
b_cells_3 = ",".join(f"F{r}" for r in chain_rows_3)
p_cells_3 = ",".join(f"H{r}" for r in chain_rows_3)
b_cells_4 = ",".join(f"F{r}" for r in chain_rows_4)
p_cells_4 = ",".join(f"H{r}" for r in chain_rows_4)

ws["A48"] = "Direct chain (water/O2/food/valves/textiles)"
ws["B48"] = f"=SUM({b_cells_3})"
ws["B48"].font = FORMULA_BLACK
ws["C48"] = f"=SUM({p_cells_3})"
ws["C48"].font = FORMULA_BLACK
ws["D48"] = f"=SUM({b_cells_3})"
ws["D48"].font = FORMULA_BLACK

ws["A49"] = "Upstream chain (smelter, mines, carbon mill)"
ws["B49"] = f"=SUM({b_cells_4})"
ws["B49"].font = FORMULA_BLACK
ws["C49"] = f"=SUM({p_cells_4})"
ws["C49"].font = FORMULA_BLACK
ws["D49"] = f"=SUM({b_cells_4})"
ws["D49"].font = FORMULA_BLACK

ws["A50"] = "Life support buildings (Water/O2/Food)"
ws["B50"] = 3
ws["B50"].font = INPUT_BLUE
ws["C50"] = "=B50*1"
ws["C50"].font = FORMULA_BLACK
ws["D50"] = "=B50"
ws["D50"].font = FORMULA_BLACK

ws["A51"] = "Storage (Silos + Tanks)"
ws["B51"] = "=B12"
ws["B51"].font = LINK_GREEN
ws["C51"] = 0
ws["C51"].font = INPUT_BLUE
ws["D51"] = "=B12"
ws["D51"].font = LINK_GREEN

ws["A52"] = "TOTAL"
ws["A52"].font = Font(name=FONT, bold=True)
ws["B52"] = "=B48+B49+B50+B51"
ws["B52"].font = Font(name=FONT, bold=True)
ws["C52"] = "=C48+C49+C50+C51"
ws["C52"].font = Font(name=FONT, bold=True)
ws["D52"] = "=D48+D49+D50+D51"
ws["D52"].font = Font(name=FONT, bold=True)

ws["A54"] = "People Capacity available"
ws["A54"].font = Font(name=FONT, bold=True)
ws["B54"] = "=B9"
ws["B54"].font = LINK_GREEN
ws["B54"].number_format = "0.0"
ws["A55"] = "Slack people (positive = healthy, negative = capacity-bound)"
ws["B55"] = "=B54-C52"
ws["B55"].font = FORMULA_BLACK
ws["B55"].number_format = "0.0;[Red]-0.0"

ws["A57"] = "Grid slots available for production"
ws["A57"].font = Font(name=FONT, bold=True)
ws["B57"] = "=B13"
ws["B57"].font = LINK_GREEN
ws["A58"] = "Slack slots (positive = healthy, negative = layout-bound)"
ws["B58"] = "=B57-D52+B50+B51"
ws["B58"].font = FORMULA_BLACK
ws["B58"].number_format = "0;[Red]-0"

cm_section(ws, 60, "6. Math-feel findings (interpret slack values)")
findings = [
    "If Slack people < 0: Comfortable colony cannot run all chains locally — must import or distribute across bodies.",
    "If Slack slots < 0: 7×7 lunar habitat is too tight for full Comfortable build-out — forces specialization or a second body.",
    "Oxygen demand at 60 pop = 10/min; Electrolyzer outputs 1 O2/cycle/60s = 1/min. ⇒ 10 Electrolyzers if local-only.",
    "  → This is the load-bearing 'Aluminum is over-demanded' (Pending #17) signal: Electrolyzer + Refinery competing for Aluminum.",
    "  → Realistic resolution at v1: import Oxygen via tanker route from a body running excess Electrolysis (multi-stop is from T0 specifically for this).",
    "Carbon Mesh demand for Textiles + Furnishings (when both active) = ~2× single-stream → Carbon Ore single-source bottleneck (Pending #16).",
    "Pressure-Valve Forge cycle 120s producing 1 PV → 0.5/min; 60-pop demand 2/min ⇒ 4 forges, 20 people. Largest single people-cost line item.",
    "If chain math is unhealthy, dial: per-pop drip rates (slowest), cycle times (medium), or Electrolyzer output qty (cleanest fix).",
]
for i, line in enumerate(findings, start=61):
    ws.cell(row=i, column=1, value=line).font = Font(name=FONT)
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

ws["A70"] = "Stage 3 playtest acceptance (math-feel)"
ws["A70"].font = SECTION_13
ws["A70"].fill = SECTION_FILL
acceptance = [
    "1. A 60-pop Comfortable lunar habitat is achievable on a 7×7 grid without exporting most needs (oxygen import is acceptable).",
    "2. People Capacity slack at full Comfortable build-out: 5–25 (not 0, not 50) — capacity is a real lever.",
    "3. Slot slack: 1–8 — layout-decision floor (R72) is real.",
    "4. Adjacency at 15–35% feels load-bearing on a 4×4 NEA (verify against the Adjacency_Pairs sheet).",
]
for i, line in enumerate(acceptance, start=71):
    ws.cell(row=i, column=1, value=line).font = Font(name=FONT)
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

# =============================================================================
out = "BALANCE.xlsx"
wb.save(out)
print(f"Wrote {out}")
